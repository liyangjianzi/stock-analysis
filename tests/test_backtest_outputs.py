"""Test backtest output builders (charts and reports).

Fully offline: fixtures are synthetic OHLCV data.
"""
from __future__ import annotations

import openpyxl
import pandas as pd

from stockanalysis import charts
from stockanalysis.backtest import build_results_from_prices
from stockanalysis.outputs.backtest_excel import write_backtest_workbook


def test_build_backtest_report_returns_figure(uptrend_ohlcv, downtrend_ohlcv):
    res = build_results_from_prices({"UP": uptrend_ohlcv, "DOWN": downtrend_ohlcv},
                                    horizons=("1m", "3m"), max_hold="1m")
    fig = charts.build_backtest_report(res)
    assert fig is not None
    assert len(fig.data) >= 1            # at least the equity curve


def test_build_backtest_report_flags_composite(uptrend_ohlcv):
    res = build_results_from_prices({"UP": uptrend_ohlcv}, mode="composite",
                                    fundamental_scores={"UP": 6}, horizons=("1m",),
                                    max_hold="1m")
    fig = charts.build_backtest_report(res)
    assert "COMPOSITE" in fig.layout.title.text.upper()


def test_write_backtest_workbook(tmp_path, uptrend_ohlcv, downtrend_ohlcv):
    res = build_results_from_prices({"UP": uptrend_ohlcv, "DOWN": downtrend_ohlcv},
                                    horizons=("1m", "3m"), max_hold="1m")
    path = tmp_path / "bt.xlsx"
    out = write_backtest_workbook(res, path)

    wb = openpyxl.load_workbook(out)
    assert "Backtest Summary" in wb.sheetnames
    assert "Event Study" in wb.sheetnames
