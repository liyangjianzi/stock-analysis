"""Tests for outputs.ExcelExporter — fully offline via openpyxl."""
from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from stockanalysis.outputs import get_exporter
from stockanalysis.outputs.excel import ExcelExporter


def _signal_matrix():
    return pd.DataFrame({
        "Ticker": ["AAPL", "JPM"],
        "Sector": ["Technology", "Financials"],
        "Final Action Signal": ["Buy", "Hold"],
    })


def _full_matrix():
    """A matrix exercising every styled column (action, posture, score scales)."""
    return pd.DataFrame({
        "Ticker": ["AAPL", "JPM", "PFE"],
        "Sector": ["Technology", "Financials", "Healthcare"],
        "Fundamental Score": [6, 4, 2],
        "Technical Posture": ["Bullish", "Neutral", "Bearish"],
        "Tech Score": [5, 3, 0],
        "Composite": [0.800, 0.500, 0.250],
        "Final Action Signal": ["Buy", "Hold", "Watch"],
    })


def _col(ws, header):
    """Column letter whose row-1 header == ``header``."""
    return next(c.column_letter for c in ws[1] if c.value == header)


def _export(tmp_path, matrix):
    out_path = tmp_path / "out.xlsx"
    ExcelExporter(path=str(out_path)).export(matrix)
    return load_workbook(out_path)["Signal Matrix"]


def test_export_writes_two_sheet_workbook(tmp_path, make_screened):
    out_path = tmp_path / "out.xlsx"
    dest = ExcelExporter(path=str(out_path)).export(
        _signal_matrix(), make_screened({"AAPL": 6})
    )

    assert dest == str(out_path)
    assert out_path.exists()
    sheets = pd.read_excel(out_path, sheet_name=None)
    assert set(sheets) == {"Signal Matrix", "Fundamentals"}
    assert sheets["Signal Matrix"]["Ticker"].tolist() == ["AAPL", "JPM"]


def test_export_without_screened_writes_single_sheet(tmp_path):
    out_path = tmp_path / "out.xlsx"
    ExcelExporter(path=str(out_path)).export(_signal_matrix())
    sheets = pd.read_excel(out_path, sheet_name=None)
    assert set(sheets) == {"Signal Matrix"}


def test_empty_matrix_writes_nothing(tmp_path):
    out_path = tmp_path / "out.xlsx"
    dest = ExcelExporter(path=str(out_path)).export(pd.DataFrame())
    assert dest == ""
    assert not out_path.exists()


def test_factory_returns_excel_exporter():
    assert isinstance(get_exporter("excel"), ExcelExporter)


# --- styling -------------------------------------------------------------------

def test_header_is_styled_and_frozen(tmp_path):
    ws = _export(tmp_path, _full_matrix())
    assert ws.freeze_panes == "A2"           # header row frozen
    assert ws.auto_filter.ref is not None    # filter dropdowns present
    assert ws["A1"].font.bold
    assert ws["A1"].fill.fgColor.rgb.endswith("1F3864")  # dark header fill


def test_action_column_is_color_coded(tmp_path):
    ws = _export(tmp_path, _full_matrix())
    col = _col(ws, "Final Action Signal")
    assert ws[f"{col}2"].value == "Buy"
    assert ws[f"{col}2"].fill.fgColor.rgb.endswith("B7E1CD")   # Buy -> green
    assert ws[f"{col}4"].value == "Watch"
    assert ws[f"{col}4"].fill.fgColor.rgb.endswith("D9D9D9")   # Watch -> gray


def test_posture_badges_are_colored(tmp_path):
    ws = _export(tmp_path, _full_matrix())
    col = _col(ws, "Technical Posture")
    assert ws[f"{col}2"].fill.fgColor.rgb.endswith("B7E1CD")   # Bullish -> green
    assert ws[f"{col}4"].fill.fgColor.rgb.endswith("F4C7C3")   # Bearish -> red


def test_score_columns_have_color_scales(tmp_path):
    ws = _export(tmp_path, _full_matrix())
    assert len(list(ws.conditional_formatting)) >= 1  # heatmap rules added


def test_styling_skips_absent_columns_without_raising(tmp_path):
    # No Technical Posture / score columns — must not raise, action still colored.
    bare = pd.DataFrame({"Ticker": ["AAPL"], "Final Action Signal": ["Buy"]})
    ws = _export(tmp_path, bare)
    col = _col(ws, "Final Action Signal")
    assert ws[f"{col}2"].fill.fgColor.rgb.endswith("B7E1CD")
