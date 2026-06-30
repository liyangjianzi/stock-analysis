"""Excel exporter (openpyxl). Works with the core dependencies — no extras.

Writes a polished, Google-Sheets-style workbook: a styled + frozen header,
banded rows, auto-fit columns, a color-coded ``Final Action Signal`` column,
``Technical Posture`` badges, and heatmap color-scales on the score columns.
Styling is layered on top of the plain values, so the data round-trips exactly.
"""
from __future__ import annotations

import logging

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .base import FUNDAMENTALS_SHEET, SIGNAL_MATRIX_SHEET, Exporter

log = logging.getLogger(__name__)

# --- Palette (Google-Sheets-like) ---------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")          # dark slate
HEADER_FONT = Font(bold=True, color="FFFFFF")
BAND_FILL = PatternFill("solid", fgColor="F3F3F3")            # even-row shading

ACTION_FILLS = {
    "Buy":   PatternFill("solid", fgColor="B7E1CD"),          # green
    "Hold":  PatternFill("solid", fgColor="FCE8B2"),          # amber
    "Watch": PatternFill("solid", fgColor="D9D9D9"),          # gray
}
POSTURE_FILLS = {
    "Bullish": PatternFill("solid", fgColor="B7E1CD"),
    "Neutral": PatternFill("solid", fgColor="FCE8B2"),
    "Bearish": PatternFill("solid", fgColor="F4C7C3"),        # red
}
# Classic 3-color heatmap: low=red, mid=yellow, high=green.
_SCORE_SCALE = dict(
    start_type="min", start_color="F8696B",
    mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max", end_color="63BE7B",
)
SCORE_COLUMNS = ("Fundamental Score", "Tech Score", "Composite")

_MAX_COL_WIDTH = 40  # cap so a long sector name doesn't blow out the layout


def _header_columns(ws) -> dict[str, str]:
    """Map each header label in row 1 to its column letter (by name, so coloring
    is robust to column order and silently skips columns that aren't present)."""
    return {
        cell.value: get_column_letter(cell.column)
        for cell in ws[1]
        if cell.value is not None
    }


def _style_base(ws) -> None:
    """Header styling, frozen header, autofilter, banded rows, auto widths.
    Applied to every sheet."""
    n_rows, n_cols = ws.max_row, ws.max_column
    if n_rows < 1 or n_cols < 1:
        return

    # Header row.
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Banded data rows (even rows shaded).
    for row in range(2, n_rows + 1):
        if row % 2 == 0:
            for col in range(1, n_cols + 1):
                ws.cell(row=row, column=col).fill = BAND_FILL

    # Auto column widths from the longest stringified value (header + cells).
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        longest = max(
            (len(str(ws.cell(row=r, column=col).value))
             for r in range(1, n_rows + 1)
             if ws.cell(row=r, column=col).value is not None),
            default=0,
        )
        ws.column_dimensions[letter].width = min(longest + 2, _MAX_COL_WIDTH)


def _style_signal_matrix(ws) -> None:
    """Signal-Matrix-only coloring: action fills, posture badges, score scales.
    Runs after _style_base; guards on column presence and ≥1 data row."""
    n_rows = ws.max_row
    if n_rows < 2:
        return
    cols = _header_columns(ws)

    # Semantic solid fills for the categorical columns (overwrite banding).
    for col_name, fills in (("Final Action Signal", ACTION_FILLS),
                            ("Technical Posture", POSTURE_FILLS)):
        letter = cols.get(col_name)
        if not letter:
            continue
        for row in range(2, n_rows + 1):
            cell = ws[f"{letter}{row}"]
            fill = fills.get(cell.value)
            if fill is not None:
                cell.fill = fill

    # Heatmap color-scales on the numeric score columns.
    for col_name in SCORE_COLUMNS:
        letter = cols.get(col_name)
        if not letter:
            continue
        rng = f"{letter}2:{letter}{n_rows}"
        ws.conditional_formatting.add(rng, ColorScaleRule(**_SCORE_SCALE))
        if col_name == "Composite":
            for row in range(2, n_rows + 1):
                ws[f"{letter}{row}"].number_format = "0.000"


class ExcelExporter(Exporter):
    """Write the signal matrix + screener detail to a styled two-sheet workbook."""

    def __init__(self, path: str = "signal_matrix.xlsx"):
        self.path = path

    def export(self, signal_matrix: pd.DataFrame,
               screened_df: pd.DataFrame | None = None, **meta) -> str:
        if signal_matrix is None or signal_matrix.empty:
            log.warning("Nothing to export — signal matrix is empty.")
            return ""
        with pd.ExcelWriter(self.path, engine="openpyxl") as xl:
            signal_matrix.to_excel(xl, sheet_name=SIGNAL_MATRIX_SHEET, index=False)
            if screened_df is not None and not screened_df.empty:
                screened_df.reset_index().to_excel(xl, sheet_name=FUNDAMENTALS_SHEET, index=False)

            # Layer styling on top of the written values.
            for name, ws in xl.sheets.items():
                _style_base(ws)
            _style_signal_matrix(xl.sheets[SIGNAL_MATRIX_SHEET])
        log.info("Exported results to '%s' (sheets: %s, %s).",
                 self.path, SIGNAL_MATRIX_SHEET, FUNDAMENTALS_SHEET)
        return self.path
